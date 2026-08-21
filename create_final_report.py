import pandas as pd
import numpy as np
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

files = [
    ('DATA SANTRI/25 06  2026_DATA SISWA KELAS 1 GABUNGAN 26-27ds.xlsx', 'Kelas 1'),
    ('DATA SANTRI/16072026_Terbaru DATA SISWA KELAS 2 GABUNGAN 26-27.xlsx', 'Kelas 2'),
    ('DATA SANTRI/18072026_Terbaru DATA SISWA KELAS 3 GABUNGAN 26-27ds.xlsx', 'Kelas 3'),
    ('DATA SANTRI/03 08 2026_DATA SISWA KELAS 4 GABUNGAN 26-27ds.xlsx', 'Kelas 4'),
    ('DATA SANTRI/20 07 2026_DATA SISWA KELAS 5 GABUNGAN 26-27ds.xlsx', 'Kelas 5'),
    ('DATA SANTRI/30 06 2026_DATA SISWA KELAS 6 GABUNGAN 26-27ds.xlsx', 'Kelas 6'),
]

all_students = []

for fpath, label in files:
    df = pd.read_excel(fpath)
    for i, row in df.iterrows():
        row_data = row.tolist()
        n_cols = len(row_data)

        if n_cols == 62:
            data = {
                'nama': row_data[4], 'nis': row_data[2], 'nisn': row_data[3],
                'jk': row_data[11], 'tempat': row_data[12], 'tgl_lahir': row_data[14],
                'ayah': row_data[27], 'ibu': row_data[34], 'alamat': row_data[18],
                'asal_sekolah': row_data[24], 'desa': row_data[20],
                'kecamatan': row_data[21], 'kota': row_data[22],
                'anak_ke': row_data[8], 'jml_saudara': row_data[9], 'telp_ayah': row_data[32],
                'tingkat': label, 'file': fpath,
            }
        else:
            data = {
                'nama': row_data[5], 'nis': row_data[3], 'nisn': row_data[4],
                'jk': row_data[12], 'tempat': row_data[13], 'tgl_lahir': row_data[15],
                'ayah': row_data[28], 'ibu': row_data[35], 'alamat': row_data[19],
                'asal_sekolah': row_data[25], 'desa': row_data[21],
                'kecamatan': row_data[22], 'kota': row_data[23],
                'anak_ke': row_data[9], 'jml_saudara': row_data[10], 'telp_ayah': row_data[33],
                'tingkat': label, 'file': fpath,
            }
        all_students.append(data)

def norm(s):
    if s is None or pd.isna(s): return ''
    return str(s).lower().strip()

# Styles
header_font = Font(bold=True, color="FFFFFF")
border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
BLUE_FILL = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")

wb = Workbook()

# =============================================
# SHEET 1: DATA DOBEL (YANG HARUS DIHAPUS)
# =============================================
ws1 = wb.active
ws1.title = "1-Dobel-HAPUS"

ws1.cell(row=1, column=1, value="DATA DOBEL - HARUS DIHAPUS").font = Font(bold=True, size=14, color="C00000")
ws1.cell(row=2, column=1, value="2 NISN ini muncul 2x di kelas yang sama. Hapus SALAH SATU.").font = Font(italic=True)

headers = ["NISN", "Nama", "Kelas", "JK", "Tempat Lahir", "Tanggal Lahir", "Ayah", "Ibu", "HAPUS?"]
for col, h in enumerate(headers, 1):
    cell = ws1.cell(row=4, column=col, value=h)
    cell.font = header_font
    cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    cell.border = border

# Find doubles
nisn_groups = defaultdict(list)
for s in all_students:
    nisn = str(s['nisn']) if s['nisn'] and not pd.isna(s['nisn']) else ''
    if nisn:
        nisn_groups[nisn].append(s)

row_num = 5
for nisn, students in nisn_groups.items():
    if len(students) > 1:
        for s in students:
            tgl = str(s['tgl_lahir'])[:10] if pd.notna(s['tgl_lahir']) else ''
            data = [nisn, s['nama'], s['tingkat'], s['jk'], s['tempat'], tgl, s['ayah'], s['ibu'], 'HAPUS']
            for col, val in enumerate(data, 1):
                cell = ws1.cell(row=row_num, column=col, value=val)
                cell.border = border
                cell.fill = RED_FILL
            row_num += 1

for i, w in enumerate([15, 35, 10, 10, 20, 15, 30, 30, 15], 1):
    ws1.column_dimensions[chr(64+i)].width = w

print(f"Sheet 1: {row_num-5} entry dobel")

# =============================================
# SHEET 2: KAKAK BERADIK TERVERIFIKASI (HASIL TRIANGULASI)
# =============================================
ws2 = wb.create_sheet("2-Siblings-Verified")

ws2.cell(row=1, column=1, value="KAKAK BERADIK TERVERIFIKASI").font = Font(bold=True, size=14)
ws2.cell(row=2, column=1, value="Diverifikasi berdasarkan triangulasi: Ayah+Ibu+Alamat atau Ayah+Ibu+Telepon").font = Font(italic=True)

# Run strict triangulation
groups1 = defaultdict(list)
for s in all_students:
    ayah = norm(s['ayah'])
    ibu = norm(s['ibu'])
    alamat = norm(s['alamat'])[:60] if s['alamat'] else ''
    if ayah and ibu and alamat and ayah != 'nan' and ibu != 'nan' and alamat != 'nan':
        groups1[(ayah, ibu, alamat)].append(s)
siblings1 = {k: v for k, v in groups1.items() if len(v) > 1}

groups2 = defaultdict(list)
for s in all_students:
    ayah = norm(s['ayah'])
    ibu = norm(s['ibu'])
    telp = str(s['telp_ayah'])[:12] if s['telp_ayah'] else ''
    if ayah and ibu and telp and ayah != 'nan' and ibu != 'nan' and len(telp) >= 10:
        groups2[(ayah, ibu, telp)].append(s)
siblings2 = {k: v for k, v in groups2.items() if len(v) > 1}

headers2 = ["No", "Nama Santri", "NISN", "Kelas", "JK", "Anak ke", "TTL", "Nama Ayah", "Nama Ibu", "Alamat", "Kota", "Telp", "Hubungan"]
for col, h in enumerate(headers2, 1):
    cell = ws2.cell(row=4, column=col, value=h)
    cell.font = header_font
    cell.fill = GREEN_FILL
    cell.border = border

row_num = 5
fam_num = 0
seen_nisns = set()
all_verified = []

for groups, strat in [(siblings1, 'Alamat'), (siblings2, 'Telp')]:
    for key, students in groups.items():
        if len(students) >= 2:
            unique = [s for s in students if str(s['nisn']) not in seen_nisns]
            if len(unique) >= 2:
                fam_num += 1
                for s in unique:
                    seen_nisns.add(str(s['nisn']))
                    tgl = str(s['tgl_lahir'])[:10] if pd.notna(s['tgl_lahir']) else ''
                    anak = int(s['anak_ke']) if pd.notna(s['anak_ke']) else ''
                    data = [fam_num, s['nama'], s['nisn'], s['tingkat'], s['jk'], anak,
                           f"{s['tempat']}, {tgl}" if s['tempat'] else tgl,
                           s['ayah'], s['ibu'], str(s['alamat'])[:50] if s['alamat'] else '',
                           s['kota'], str(s['telp_ayah'])[:15] if s['telp_ayah'] else '',
                           strat]
                    for col, val in enumerate(data, 1):
                        cell = ws2.cell(row=row_num, column=col, value=val)
                        cell.border = border
                        cell.fill = GREEN_FILL
                    row_num += 1
                    all_verified.append(s)

for i, w in enumerate([6, 30, 12, 8, 8, 8, 25, 25, 25, 50, 15, 15, 15], 1):
    ws2.column_dimensions[chr(64+i)].width = w

print(f"Sheet 2: {fam_num} keluarga siblings ({row_num-5} siswa)")

# =============================================
# SHEET 3: KEMBAR (DITERVERIFIKASI MANUAL)
# =============================================
ws3 = wb.create_sheet("3-Kembar")

ws3.cell(row=1, column=1, value="KEMBAR").font = Font(bold=True, size=14, color="7030A0")
ws3.cell(row=2, column=1, value="Daffa Adya Rachmad & Daffi Adya Rachmad - Kembar").font = Font(italic=True)

headers3 = ["No", "Nama", "NISN", "Kelas", "JK", "TTL", "Ayah", "Ibu", "Alamat", "Status"]
for col, h in enumerate(headers3, 1):
    cell = ws3.cell(row=4, column=col, value=h)
    cell.font = header_font
    cell.fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
    cell.border = border

# Find kembar
kembar_names = ['daffa adya rachmad', 'daffi adya rachmad']
row_num = 5
for s in all_students:
    if norm(s['nama']) in kembar_names:
        tgl = str(s['tgl_lahir'])[:10] if pd.notna(s['tgl_lahir']) else ''
        data = [1, s['nama'], s['nisn'], s['tingkat'], s['jk'], tgl, s['ayah'], s['ibu'], s['alamat'], 'KEMBAR']
        for col, val in enumerate(data, 1):
            cell = ws3.cell(row=row_num, column=col, value=val)
            cell.border = border
            cell.fill = PatternFill(start_color="E4DFEC", end_color="E4DFEC", fill_type="solid")
        row_num += 1

for i, w in enumerate([6, 30, 12, 8, 8, 20, 25, 25, 50, 15], 1):
    ws3.column_dimensions[chr(64+i)].width = w

# =============================================
# SHEET 4: ADIK KAKAK (DITERVERIFIKASI MANUAL)
# =============================================
ws4 = wb.create_sheet("4-AdikKakak")

ws4.cell(row=1, column=1, value="ADIK KAKAK").font = Font(bold=True, size=14, color="0070C0")
ws4.cell(row=2, column=1, value="Firaas Izzulhaq & Zaidan Mumtaz - Adik Kakak").font = Font(italic=True)

headers4 = ["No", "Nama", "NISN", "Kelas", "JK", "Anak ke", "TTL", "Ayah", "Ibu", "Alamat", "Status"]
for col, h in enumerate(headers4, 1):
    cell = ws4.cell(row=4, column=col, value=h)
    cell.font = header_font
    cell.fill = BLUE_FILL
    cell.border = border

# Find siblings
sibling_names = ['firaas izzulhaq', 'zaidan mumtaz']
row_num = 5
for s in all_students:
    if norm(s['nama']) in sibling_names:
        tgl = str(s['tgl_lahir'])[:10] if pd.notna(s['tgl_lahir']) else ''
        anak = int(s['anak_ke']) if pd.notna(s['anak_ke']) else ''
        status = 'KAKAK' if anak == 1 else 'ADIK'
        data = [1, s['nama'], s['nisn'], s['tingkat'], s['jk'], anak,
               f"{s['tempat']}, {tgl}" if s['tempat'] else tgl,
               s['ayah'], s['ibu'], str(s['alamat'])[:50] if s['alamat'] else '', status]
        for col, val in enumerate(data, 1):
            cell = ws4.cell(row=row_num, column=col, value=val)
            cell.border = border
            cell.fill = PatternFill(start_color="DEEBF7", end_color="DEEBF7", fill_type="solid")
        row_num += 1

for i, w in enumerate([6, 30, 12, 8, 8, 8, 25, 25, 25, 50, 15], 1):
    ws4.column_dimensions[chr(64+i)].width = w

# =============================================
# SHEET 5: BUKAN KAKAK BERADIK (ALAMAT SAMA TAPI ORTU BERBEDA)
# =============================================
ws5 = wb.create_sheet("5-BukanSiblings")

ws5.cell(row=1, column=1, value="BUKAN KAKAK BERADIK - ALAMAT SAMA TAPI ORANG TUA BERBEDA").font = Font(bold=True, size=14, color="666666")
ws5.cell(row=2, column=1, value="7 pasang ini tinggal di alamat sama tapi orang tuanya berbeda - BUKAN bersaudara").font = Font(italic=True)

headers5 = ["Nama 1", "NISN 1", "Kelas 1", "Nama Ayah 1", "Nama 2", "NISN 2", "Kelas 2", "Nama Ayah 2", "Alamat", "Status"]
for col, h in enumerate(headers5, 1):
    cell = ws5.cell(row=4, column=col, value=h)
    cell.font = header_font
    cell.fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
    cell.border = border

# Based on user verification
confirmed_not_siblings = [
    ('hafizh adiyatma nugraha', '121537258', 'kelas 2', 'akhmad agung nugroho', 'kevin wahyu ilahi', '124560445', 'kelas 2', 'tetep sutikno', 'blok mawar...', 'BEDA KELUARGA'),
    ('muhammad fikri adnan', '115041337', 'kelas 5', 'samsyul bukhari', 'muhammad fayyadh irsyad', '95045610', 'kelas 5', 'syamsul bukhari', 'wirobrajan...', 'BEDA KELUARGA'),
    ('muhammad rafi\' \'athaullah', '87026465', 'kelas 6', 'wahyudiyono', 'azmi akrami maha', '94806637', 'kelas 6', 'h. guntar maha', 'jatisawit asri...', 'BEDA KELUARGA'),
    ('m. irham ismail', '95456465', 'kelas 6', '?', 'muhammad fajri ramadhan', '89781780', 'kelas 6', '?', 'giwangan asri...', 'BEDA KELUARGA'),
    ('fairuz fawwazul akmal', '87146774', 'kelas 6', '?', 'naufal syamil adz dzaki', '97419393', 'kelas 6', '?', 'kol sugiono...', 'BEDA KELUARGA'),
]

row_num = 5
for data in confirmed_not_siblings:
    for col, val in enumerate(data, 1):
        cell = ws5.cell(row=row_num, column=col, value=val)
        cell.border = border
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    row_num += 1

for i, w in enumerate([30, 12, 10, 25, 30, 12, 10, 25, 40, 15], 1):
    ws5.column_dimensions[chr(64+i)].width = w

# =============================================
# SHEET 6: RINGKASAN AKHIR
# =============================================
ws6 = wb.create_sheet("6-Ringkasan")

summary = [
    ["LAPORAN ANALISIS DATA SANTRI - FINAL", ""],
    ["Generated: 21 Agustus 2026", ""],
    ["", ""],
    ["STATISTIK TOTAL", ""],
    ["Total Santri", len(all_students), ""],
    ["", ""],
    ["TEMUAN & AKSI", "JUMLAH", "AKSI"],
    ["1. NISN Dobel (double entry)", sum(1 for s in nisn_groups.values() if len(s) > 1), "HAPUS SALAH SATU"],
    ["2. Kakak Beradik (triangulasi)", fam_num, "VERIFIED - MASUKKAN KE APP"],
    ["3. Kembar", 1, "VERIFIED - MASUKKAN KE APP"],
    ["4. Adik Kakak (manual)", 1, "VERIFIED - MASUKKAN KE APP"],
    ["5. Alamat Sama Tapi Bukan Siblings", 5, "OK - TIDAK ADA AKSI"],
    ["", ""],
    ["DETAIL SIBLINGS:", "", ""],
    ["KELUARGA 1:", "Abdul Somad & Siti Hamdanah", ""],
    ["  - Azka Aiman Najwaan (Kelas 2)", "", ""],
    ["  - Muhammad Thoriq Shidqi Sabiq (Kelas 6)", "", ""],
    ["", "", ""],
    ["KELUARGA 2:", "Puguh Bagus Narimo & Sri Utami", ""],
    ["  - Aryatama Febrian Danendra (Kelas 1)", "", ""],
    ["  - Adyatma Abdi Danendra (Kelas 3)", "", ""],
    ["", "", ""],
    ["KELUARGA 3:", "Bachtiar DWI Kurniawan & Mamik Fatayatur Rohmah", ""],
    ["  - Ahmad Mumtaz Dhiya El Haq (Kelas 1)", "", ""],
    ["  - Ahmad Ghozi El Muntazhor (Kelas 3)", "", ""],
    ["", "", ""],
    ["KELUARGA 4:", "Ahmad Arif Fadlil & Nur Hikmah", ""],
    ["  - Ahmad Amirul A'zam (Kelas 1)", "", ""],
    ["  - Ahmad Dahlan Asy'ari (Kelas 3)", "", ""],
    ["", "", ""],
    ["KEMBAR:", "Rochmad & Rohmat", ""],
    ["  - Daffa Adya Rachmad (Kelas 6)", "", ""],
    ["  - Daffi Adya Rachmad (Kelas 6)", "", ""],
    ["", "", ""],
    ["ADIK KAKAK:", "Ihsan Effendi", ""],
    ["  - Firaas Izzulhaq (Anak ke 3, Kelas 3)", "", ""],
    ["  - Zaidan Mumtaz (Anak ke 2, Kelas 5)", "", ""],
]

for row_data in summary:
    ws6.append(row_data)

ws6.column_dimensions['A'].width = 45
ws6.column_dimensions['B'].width = 40

# Save
output_path = "DATA SANTRI/LAPORAN_FINAL_SIBLINGS.xlsx"
wb.save(output_path)
print(f"\nLaporan FINAL disimpan: {output_path}")
print(f"\nSUMMARY:")
print(f"  - Total Santri: {len(all_students)}")
print(f"  - NISN Dobel: {sum(1 for s in nisn_groups.values() if len(s) > 1)} (HARUS HAPUS)")
print(f"  - Kakak Beradik (Triangulasi): {fam_num} keluarga")
print(f"  - Kembar: 1 pasang (Daffa & Daffi)")
print(f"  - Adik Kakak: 1 pasang (Firaas & Zaidan)")
print(f"  - Bukan Siblings: 5 pasang (alamat sama, ortu berbeda)")
