import pandas as pd
import numpy as np
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

files = [
    ('DATA SANTRI/25 06  2026_DATA SISWA KELAS 1 GABUNGAN 26-27ds.xlsx', 'Kelas 1'),
    ('DATA SANTRI/16072026_Terbaru DATA SISWA KELAS 2 GABUNGAN 26-27.xlsx', 'Kelas 2'),
    ('DATA SANTRI/18072026_Terbaru DATA SISWA KELAS 3 GABUNGAN 26-27ds.xlsx', 'Kelas 3'),
    ('DATA SANTRI/03 08 2026_DATA SISWA KELAS 4 GABUNGAN 26-27ds.xlsx', 'Kelas 4'),
    ('DATA SANTRI/20 07 2026_DATA SISWA KELAS 5 GABUNGAN 26-27ds.xlsx', 'Kelas 5'),
    ('DATA SANTRI/30 06 2026_DATA SISWA KELAS 6 GABUNGAN 26-27ds.xlsx', 'Kelas 6'),
]

all_students = []

for fpath, label in files:
    df = pd.read_excel(fpath)

    for i, row in df.iterrows():
        row_data = row.tolist()
        n_cols = len(row_data)

        if n_cols == 62:
            nama = row_data[4]
            nis = row_data[2]
            nisn = row_data[3]
            jk = row_data[11]
            tempat = row_data[12]
            tgl_lahir = row_data[14]
            ayah = row_data[27]
            ibu = row_data[34]
            alamat = row_data[18]
            asal_sekolah = row_data[24]
            desa = row_data[20]
            kecamatan = row_data[21]
            kota = row_data[22]
            anak_ke = row_data[8]
            jml_saudara = row_data[9]
            telp_ayah = row_data[32]
        else:
            nama = row_data[5]
            nis = row_data[3]
            nisn = row_data[4]
            jk = row_data[12]
            tempat = row_data[13]
            tgl_lahir = row_data[15]
            ayah = row_data[28]
            ibu = row_data[35]
            alamat = row_data[19]
            asal_sekolah = row_data[25]
            desa = row_data[21]
            kecamatan = row_data[22]
            kota = row_data[23]
            anak_ke = row_data[9]
            jml_saudara = row_data[10]
            telp_ayah = row_data[33]

        all_students.append({
            'nama': nama, 'nis': nis, 'nisn': nisn, 'jk': jk,
            'tempat': tempat, 'tgl_lahir': tgl_lahir,
            'ayah': ayah, 'ibu': ibu, 'alamat': alamat,
            'asal_sekolah': asal_sekolah, 'desa': desa,
            'kecamatan': kecamatan, 'kota': kota,
            'anak_ke': anak_ke, 'jml_saudara': jml_saudara,
            'telp_ayah': telp_ayah,
            'tingkat': label,
        })

def norm(s):
    if s is None or pd.isna(s):
        return ''
    return str(s).lower().strip()

# TRIANGULASI
groups1 = defaultdict(list)
for s in all_students:
    ayah = norm(s['ayah'])
    ibu = norm(s['ibu'])
    alamat = norm(s['alamat'])[:50] if s['alamat'] else ''
    if ayah and ibu and alamat and ayah != 'nan' and ibu != 'nan' and alamat != 'nan':
        groups1[(ayah, ibu, alamat)].append(s)

groups2 = defaultdict(list)
for s in all_students:
    ayah = norm(s['ayah'])
    ibu = norm(s['ibu'])
    telp = str(s['telp_ayah'])[:10] if s['telp_ayah'] else ''
    if ayah and ibu and telp and ayah != 'nan' and ibu != 'nan' and len(telp) >= 8:
        groups2[(ayah, ibu, telp)].append(s)

groups3 = defaultdict(list)
for s in all_students:
    alamat = norm(s['alamat'])[:40] if s['alamat'] else ''
    anak_ke = s['anak_ke']
    if alamat and anak_ke and alamat != 'nan' and pd.notna(anak_ke):
        groups3[alamat].append(s)

def has_consecutive(students):
    anak_kes = sorted([s['anak_ke'] for s in students if pd.notna(s['anak_ke'])])
    if len(anak_kes) >= 2:
        for i in range(len(anak_kes) - 1):
            if anak_kes[i+1] - anak_kes[i] <= 2:
                return True
    return False

siblings3 = {k: v for k, v in groups3.items() if has_consecutive(v)}

groups4 = defaultdict(list)
for s in all_students:
    desa = norm(s['desa'])
    kecamatan = norm(s['kecamatan'])
    anak_ke = s['anak_ke']
    if desa and kecamatan and anak_ke and desa != 'nan' and kecamatan != 'nan' and pd.notna(anak_ke):
        key = f"{desa}|{kecamatan}"
        groups4[key].append(s)

siblings4 = {k: v for k, v in groups4.items() if has_consecutive(v)}

# Scoring
student_scores = defaultdict(set)
all_groups = [
    (groups1, 'Alamat'),
    (groups2, 'Telp'),
    (siblings3, 'AnakKe'),
    (siblings4, 'Desa')
]

for groups, strat_name in all_groups:
    for key, students in groups.items():
        for s in students:
            nisn = str(s['nisn']) if pd.notna(s['nisn']) else id(s)
            student_scores[nisn].add(strat_name)

verified_nisns = {n for n, strats in student_scores.items() if len(strats) >= 2}

# Collect unique families
unique_families = []
seen_nisns = set()

for groups, _ in all_groups:
    for key, students in groups.items():
        family_verified = sum(1 for s in students if str(s['nisn']) in verified_nisns)
        if family_verified >= 2 and len(students) >= 2:
            family_students = []
            for s in students:
                nisn = str(s['nisn']) if pd.notna(s['nisn']) else id(s)
                if nisn not in seen_nisns:
                    seen_nisns.add(nisn)
                    family_students.append(s)
            if len(family_students) >= 2:
                unique_families.append((key, family_students))

seen_sigs = set()
final_families = []
for key, students in unique_families:
    sig = tuple(sorted([str(s['nisn']) for s in students if pd.notna(s['nisn'])]))
    if sig not in seen_sigs and len(sig) >= 2:
        seen_sigs.add(sig)
        final_families.append((key, students))

print(f"Total keluarga: {len(final_families)}")

# CREATE EXCEL REPORT
wb = Workbook()
ws = wb.active
ws.title = "Kakak Beradik"

# Styles
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
alt_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
warning_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Headers
headers = ["No", "No KK", "Nama Santri", "NISN", "Kelas", "JK", "Anak ke", "Jml Saudara",
           "Tempat Lahir", "Tanggal Lahir", "Nama Ayah", "Nama Ibu", "Alamat", "Desa", "Kecamatan", "Kota", "Telp Ayah", "Keterangan"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', wrap_text=True)
    cell.border = border

# Data
row_num = 2
for fam_idx, (key, students) in enumerate(final_families, 1):
    for s in students:
        tgl = str(s['tgl_lahir'])[:10] if pd.notna(s['tgl_lahir']) else ''
        anak = int(s['anak_ke']) if pd.notna(s['anak_ke']) else ''
        jml = int(s['jml_saudara']) if pd.notna(s['jml_saudara']) else ''
        ayah = s['ayah'] if s['ayah'] else ''
        ibu = s['ibu'] if s['ibu'] else ''
        alamat = str(s['alamat'])[:100] if s['alamat'] else ''
        telp = str(s['telp_ayah'])[:20] if s['telp_ayah'] else ''
        desa = s['desa'] if s['desa'] else ''
        kecamatan = s['kecamatan'] if s['kecamatan'] else ''
        kota = s['kota'] if s['kota'] else ''

        # Determine if verified
        nisn = str(s['nisn']) if pd.notna(s['nisn']) else ''
        verifikasi = f"Terifikasi ({len(student_scores.get(nisn, set()))} criteria)" if nisn in verified_nisns else "Low confidence"

        data = [fam_idx, '', s['nama'], s['nisn'], s['tingkat'], s['jk'], anak, jml,
                s['tempat'], tgl, ayah, ibu, alamat, desa, kecamatan, kota, telp, verifikasi]

        fill = alt_fill if fam_idx % 2 == 0 else None
        if verifikasi == "Low confidence":
            fill = warning_fill

        for col, val in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.border = border
            if fill:
                cell.fill = fill

        row_num += 1

# Adjust column widths
col_widths = [6, 6, 30, 12, 8, 6, 8, 8, 15, 12, 25, 25, 50, 15, 15, 15, 15, 20]
for i, width in enumerate(col_widths, 1):
    ws.column_dimensions[chr(64 + i) if i <= 26 else 'A' + chr(64 + i - 26)].width = width

# =============================================
# SHEET 2: DATA DOBEL
# =============================================
ws2 = wb.create_sheet("Data Dobel")

headers2 = ["NISN", "Nama", "Kelas", "JK", "Tempat Lahir", "Tanggal Lahir", "Aksi"]
for col, header in enumerate(headers2, 1):
    cell = ws2.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    cell.border = border

# Find duplicates
nisn_groups = defaultdict(list)
for s in all_students:
    nisn = str(s['nisn']) if pd.notna(s['nisn']) else ''
    if nisn:
        nisn_groups[nisn].append(s)

row_num = 2
for nisn, students in nisn_groups.items():
    if len(students) > 1:
        for s in students:
            tgl = str(s['tgl_lahir'])[:10] if pd.notna(s['tgl_lahir']) else ''
            data = [nisn, s['nama'], s['tingkat'], s['jk'], s['tempat'], tgl, 'HAPUS']
            for col, val in enumerate(data, 1):
                cell = ws2.cell(row=row_num, column=col, value=val)
                cell.border = border
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            row_num += 1

# =============================================
# SHEET 3: RINGKASAN
# =============================================
ws3 = wb.create_sheet("Ringkasan")

summary_data = [
    ["LAPORAN ANALISIS DATA SANTRI", ""],
    ["", ""],
    ["Total Santri", len(all_students)],
    ["Total Keluarga Kakak Beradik", len(final_families)],
    ["", ""],
    ["DATA DOBEL (NISN sama)", ""],
    ["Jumlah NISN dobel", sum(1 for s in nisn_groups.values() if len(s) > 1)],
    ["Jumlah entry dobel", sum(len(s) - 1 for s in nisn_groups.values() if len(s) > 1)],
    ["", ""],
    ["DATA PER KELAS", ""],
]

class_counts = defaultdict(int)
for s in all_students:
    class_counts[s['tingkat']] += 1

for kelas in ['Kelas 1', 'Kelas 2', 'Kelas 3', 'Kelas 4', 'Kelas 5', 'Kelas 6']:
    summary_data.append([kelas, class_counts[kelas]])

for row_data in summary_data:
    ws3.append(row_data)

ws3.column_dimensions['A'].width = 35
ws3.column_dimensions['B'].width = 15

# Save
output_path = "DATA SANTRI/LAPORAN_KAKAK_BERADIK.xlsx"
wb.save(output_path)
print(f"Laporan disimpan ke: {output_path}")

# =============================================
# SHEET 4: ANALISIS LANJUTAN
# =============================================
ws4 = wb.create_sheet("Analisis Lanjutan")

# Find potential typos in names
name_groups = defaultdict(list)
for s in all_students:
    nama = str(s['nama']).lower().strip() if s['nama'] else ''
    if nama:
        words = nama.split()
        key = ' '.join(words[:3]) if len(words) >= 3 else nama
        name_groups[key].append(s)

headers4 = ["Nama Similar", "Jumlah", "NISN 1", "Nama 1", "NISN 2", "Nama 2", "Status"]
for col, header in enumerate(headers4, 1):
    cell = ws4.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    cell.border = border

row_num = 2
for name, students in sorted(name_groups.items(), key=lambda x: -len(x[1])):
    if len(students) > 1:
        # Check if they are actually the same person
        nisns = set(str(s['nisn']) for s in students if pd.notna(s['nisn']))
        if len(nisns) == 1:
            status = "DOUBLE ENTRY - HAPUS"
            fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        elif len(nisns) == len(students):
            # Different NISNs - check TTL
            ttls = set(str(s['tgl_lahir'])[:10] if pd.notna(s['tgl_lahir']) else '' for s in students)
            if len(ttls) == 1:
                status = "MUNGKIN ORANG SAMA (NISN beda)"
                fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            else:
                status = "NAMA KEBETULAN SAMA"
                fill = alt_fill
        else:
            status = "PERLU DICEK"
            fill = warning_fill

        for col, val in enumerate([name, len(students),
                                    students[0]['nisn'] if pd.notna(students[0]['nisn']) else '',
                                    students[0]['nama'],
                                    students[1]['nisn'] if pd.notna(students[1]['nisn']) else '' if len(students) > 1 else '',
                                    students[1]['nama'] if len(students) > 1 else '',
                                    status], 1):
            cell = ws4.cell(row=row_num, column=col, value=val)
            cell.border = border
            cell.fill = fill
        row_num += 1

for i in range(1, 8):
    ws4.column_dimensions[chr(64 + i)].width = [15, 10, 15, 35, 15, 35, 30][i-1]

# Re-save with additional sheet
wb.save(output_path)
print(f"Laporan lengkap disimpan ke: {output_path}")
