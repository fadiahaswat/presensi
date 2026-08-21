import pandas as pd
import numpy as np
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

files = [
    ('DATA SANTRI/25 06  2026_DATA SISWA KELAS 1 GABUNGAN 26-27ds.xlsx', 'Kelas 1'),
    ('DATA SANTRI/16072026_Terbaru DATA SISWA KELAS 2 GABUNGAN 26-27.xlsx', 'Kelas 2'),
    ('DATA SANTRI/18072026_Terbaru DATA SISWA KELAS 3 GABUNGAN 26-27ds.xlsx', 'Kelas 3'),
    ('DATA SANTRI/03 08 2026_DATA SISWA KELAS 4 GABUNGAN 26-27ds.xlsx', 'Kelas 4'),
    ('DATA SANTRI/20 07 2026_DATA SISWA KELAS 5 GABUNGAN 26-27ds.xlsx', 'Kelas 5'),
    ('DATA SANTRI/30 06 2026_DATA SISWA KELAS 6 GABUNGAN 26-27ds.xlsx', 'Kelas 6'),
]

all_students = []

for fpath, label in files:
    df = pd.read_excel(fpath)
    for i, row in df.iterrows():
        row_data = row.tolist()
        n_cols = len(row_data)

        if n_cols == 62:
            data = {
                'nama': row_data[4], 'nis': row_data[2], 'nisn': row_data[3],
                'jk': row_data[11], 'tempat': row_data[12], 'tgl_lahir': row_data[14],
                'ayah': row_data[27], 'ibu': row_data[34], 'alamat': row_data[18],
                'asal_sekolah': row_data[24], 'desa': row_data[20],
                'kecamatan': row_data[21], 'kota': row_data[22],
                'anak_ke': row_data[8], 'jml_saudara': row_data[9], 'telp_ayah': row_data[32],
                'tingkat': label,
            }
        else:
            data = {
                'nama': row_data[5], 'nis': row_data[3], 'nisn': row_data[4],
                'jk': row_data[12], 'tempat': row_data[13], 'tgl_lahir': row_data[15],
                'ayah': row_data[28], 'ibu': row_data[35], 'alamat': row_data[19],
                'asal_sekolah': row_data[25], 'desa': row_data[21],
                'kecamatan': row_data[22], 'kota': row_data[23],
                'anak_ke': row_data[9], 'jml_saudara': row_data[10], 'telp_ayah': row_data[33],
                'tingkat': label,
            }
        all_students.append(data)

def norm(s):
    if s is None or pd.isna(s): return ''
    return str(s).lower().strip()

print("="*80)
print("TRIANGULASI KAKAK BERADIK (STRICT)")
print("="*80)

# STRATEGI 1: AYAH + IBU + ALAMAT SAMA (PRESISI TINGGI)
groups1 = defaultdict(list)
for s in all_students:
    ayah = norm(s['ayah'])
    ibu = norm(s['ibu'])
    alamat = norm(s['alamat'])[:60] if s['alamat'] else ''
    if ayah and ibu and alamat and ayah != 'nan' and ibu != 'nan' and alamat != 'nan':
        groups1[(ayah, ibu, alamat)].append(s)

siblings1 = {k: v for k, v in groups1.items() if len(v) > 1}
print(f"1. Ayah + Ibu + Alamat: {len(siblings1)} keluarga")

# STRATEGI 2: AYAH + IBU + TELEPON SAMA (PRESISI TINGGI)
groups2 = defaultdict(list)
for s in all_students:
    ayah = norm(s['ayah'])
    ibu = norm(s['ibu'])
    telp = str(s['telp_ayah'])[:12] if s['telp_ayah'] else ''
    if ayah and ibu and telp and ayah != 'nan' and ibu != 'nan' and len(telp) >= 10:
        groups2[(ayah, ibu, telp)].append(s)

siblings2 = {k: v for k, v in groups2.items() if len(v) > 1}
print(f"2. Ayah + Ibu + Telepon: {len(siblings2)} keluarga")

# STRATEGI 3: AYAH + IBU SAMA + ANAK KE BERURUTAN
def has_consecutive(students):
    anak_kes = sorted([s['anak_ke'] for s in students if pd.notna(s['anak_ke'])])
    if len(anak_kes) >= 2:
        for i in range(len(anak_kes) - 1):
            if anak_kes[i+1] - anak_kes[i] <= 2:
                return True
    return False

groups3 = defaultdict(list)
for s in all_students:
    ayah = norm(s['ayah'])
    ibu = norm(s['ibu'])
    anak_ke = s['anak_ke']
    if ayah and ibu and anak_ke and ayah != 'nan' and ibu != 'nan' and pd.notna(anak_ke):
        key = (ayah, ibu)
        groups3[key].append(s)

siblings3 = {k: v for k, v in groups3.items() if len(v) > 1 and has_consecutive(v)}
print(f"3. Ayah + Ibu + Anak Ke berurutan: {len(siblings3)} keluarga")

# STRATEGI 4: ALAMAT SAMA + ANAK KE BERURUTAN
groups4 = defaultdict(list)
for s in all_students:
    alamat = norm(s['alamat'])[:50] if s['alamat'] else ''
    anak_ke = s['anak_ke']
    if alamat and anak_ke and alamat != 'nan' and pd.notna(anak_ke):
        groups4[alamat].append(s)

siblings4 = {k: v for k, v in groups4.items() if len(v) > 1 and has_consecutive(v)}
print(f"4. Alamat + Anak Ke berurutan: {len(siblings4)} keluarga")

# VERIFIKASI SILANG - Siswa harus ada di minimal 2 strategi
print("\n" + "="*80)
print("VERIFIKASI SILANG")
print("="*80)

all_strategies = [
    (siblings1, 'Ayah+Ibu+Alamat'),
    (siblings2, 'Ayah+Ibu+Telp'),
    (siblings3, 'Ayah+Ibu+AnakKe'),
    (siblings4, 'Alamat+AnakKe')
]

student_matches = defaultdict(set)
for groups, strat_name in all_strategies:
    for key, students in groups.items():
        for s in students:
            nisn = str(s['nisn']) if pd.notna(s['nisn']) else str(id(s))
            student_matches[nisn].add(strat_name)

# Siswa yang terverifikasi ada di 2+ strategi
verified_nisns = {n for n, strats in student_matches.items() if len(strats) >= 2}
print(f"Siswa terverifikasi (2+ strategi): {len(verified_nisns)}")

# Kumpulkan keluarga unik
all_family_nisns = defaultdict(list)
for groups, strat_name in all_strategies:
    for key, students in groups.items():
        verified_in_group = [s for s in students if str(s['nisn']) in verified_nisns]
        if len(verified_in_group) >= 2:
            for s in students:
                nisn = str(s['nisn']) if pd.notna(s['nisn']) else str(id(s))
                all_family_nisns[nisn].extend(verified_in_group)

# Filter keluarga yang sudah terverifikasi
final_families = []
seen_nisns = set()

for groups, strat_name in all_strategies:
    for key, students in groups.items():
        verified = [s for s in students if str(s['nisn']) in verified_nisns]
        if len(verified) >= 2:
            unique_students = []
            for s in students:
                nisn = str(s['nisn']) if pd.notna(s['nisn']) else str(id(s))
                if nisn not in seen_nisns:
                    seen_nisns.add(nisn)
                    unique_students.append(s)
            if len(unique_students) >= 2:
                final_families.append((key, unique_students, strat_name))

# Hapus duplikat berdasarkan NISN signature
seen_sigs = set()
unique_families = []
for key, students, strat in final_families:
    sig = tuple(sorted([str(s['nisn']) for s in students if pd.notna(s['nisn'])]))
    if sig not in seen_sigs and len(sig) >= 2:
        seen_sigs.add(sig)
        unique_families.append((key, students, strat))

print(f"\nTotal keluarga unik terverifikasi: {len(unique_families)}")

# =============================================
# CREATE EXCEL REPORT
# =============================================
print("\nMembuat laporan Excel...")

wb = Workbook()

# Styles
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
alt_fill = PatternFill(start_color="D6DCE5", end_color="D6DCE5", fill_type="solid")
verify_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
warning_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
danger_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# SHEET 1: KAKAK BERADIK
ws1 = wb.active
ws1.title = "Kakak Beradik"

headers = ["No", "No KK", "Nama Santri", "NISN", "Kelas", "JK", "Anak ke", "Jml Saudara",
           "Tempat Lahir", "Tanggal Lahir", "Nama Ayah", "Nama Ibu", "Alamat (60char)",
           "Desa", "Kecamatan", "Kota", "Telp Ayah", "Strategi Verifikasi", "Confidence"]

for col, header in enumerate(headers, 1):
    cell = ws1.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', wrap_text=True)
    cell.border = border

row_num = 2
for fam_idx, (key, students, strat) in enumerate(unique_families, 1):
    for s in students:
        tgl = str(s['tgl_lahir'])[:10] if pd.notna(s['tgl_lahir']) else ''
        anak = int(s['anak_ke']) if pd.notna(s['anak_ke']) else ''
        jml = int(s['jml_saudara']) if pd.notna(s['jml_saudara']) else ''

        nisn = str(s['nisn']) if pd.notna(s['nisn']) else ''
        num_strat = len(student_matches.get(nisn, set()))

        if num_strat >= 3:
            confidence = "HIGH"
            fill = verify_fill
        elif num_strat >= 2:
            confidence = "MEDIUM"
            fill = alt_fill
        else:
            confidence = "LOW"
            fill = warning_fill

        data = [
            fam_idx, '', s['nama'], s['nisn'], s['tingkat'], s['jk'], anak, jml,
            s['tempat'], tgl, s['ayah'], s['ibu'],
            str(s['alamat'])[:60] if s['alamat'] else '',
            s['desa'], s['kecamatan'], s['kota'],
            str(s['telp_ayah'])[:15] if s['telp_ayah'] else '',
            strat, confidence
        ]

        for col, val in enumerate(data, 1):
            cell = ws1.cell(row=row_num, column=col, value=val)
            cell.border = border
            cell.fill = fill
        row_num += 1

# Set column widths
widths = [6, 6, 32, 13, 8, 6, 8, 8, 15, 12, 28, 28, 60, 18, 18, 18, 15, 25, 10]
for i, w in enumerate(widths, 1):
    col_letter = chr(64 + i) if i <= 26 else 'A' + chr(64 + i - 26)
    ws1.column_dimensions[col_letter].width = w

# SHEET 2: DATA DOBEL
ws2 = wb.create_sheet("Data Dobel")

headers2 = ["NISN", "Nama", "Kelas", "JK", "Tempat Lahir", "Tanggal Lahir", "Aksi yang Direkomendasikan"]
for col, header in enumerate(headers2, 1):
    cell = ws2.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    cell.border = border

nisn_groups = defaultdict(list)
for s in all_students:
    nisn = str(s['nisn']) if pd.notna(s['nisn']) else ''
    if nisn:
        nisn_groups[nisn].append(s)

row_num = 2
for nisn, students in nisn_groups.items():
    if len(students) > 1:
        for s in students:
            tgl = str(s['tgl_lahir'])[:10] if pd.notna(s['tgl_lahir']) else ''
            data = [nisn, s['nama'], s['tingkat'], s['jk'], s['tempat'], tgl, 'HAPUS SALAH SATU']
            for col, val in enumerate(data, 1):
                cell = ws2.cell(row=row_num, column=col, value=val)
                cell.border = border
                cell.fill = danger_fill
            row_num += 1

for i, w in enumerate([15, 35, 10, 10, 20, 15, 30], 1):
    ws2.column_dimensions[chr(64 + i)].width = w

# SHEET 3: NAMA SERUPA
ws3 = wb.create_sheet("Nama Serupa")

headers3 = ["Nama Similar", "Count", "NISN", "Nama Lengkap", "Kelas", "TTL", "Ayah", "Status"]
for col, header in enumerate(headers3, 1):
    cell = ws3.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = PatternFill(start_color="538135", end_color="538135", fill_type="solid")
    cell.border = border

# Group by first 3 words of name
name_groups = defaultdict(list)
for s in all_students:
    nama = str(s['nama']).lower().strip() if s['nama'] else ''
    if nama:
        words = nama.split()
        key = ' '.join(words[:3]) if len(words) >= 3 else nama
        name_groups[key].append(s)

row_num = 2
for name, students in sorted(name_groups.items(), key=lambda x: -len(x[1])):
    if len(students) > 1:
        nisns = set(str(s['nisn']) for s in students if pd.notna(s['nisn']))
        if len(nisns) == 1:
            status = "DOUBLE ENTRY - HAPUS!"
            fill = danger_fill
        elif len(nisns) == len(students):
            ttls = set(str(s['tgl_lahir'])[:10] if pd.notna(s['tgl_lahir']) else '' for s in students)
            if len(ttls) == 1:
                status = "MUNGKIN ORANG SAMA (NISN berbeda)"
                fill = warning_fill
            else:
                status = "Nama kebetulan sama"
                fill = alt_fill
        else:
            status = "PERLU CEK MANUAL"
            fill = warning_fill

        for idx, s in enumerate(students):
            tgl = str(s['tgl_lahir'])[:10] if pd.notna(s['tgl_lahir']) else ''
            data = [name, len(students), s['nisn'], s['nama'], s['tingkat'], tgl, s['ayah'], status]
            for col, val in enumerate(data, 1):
                cell = ws3.cell(row=row_num, column=col, value=val)
                cell.border = border
                cell.fill = fill
            row_num += 1

for i, w in enumerate([30, 8, 15, 35, 10, 15, 25, 30], 1):
    ws3.column_dimensions[chr(64 + i)].width = w

# SHEET 4: RINGKASAN
ws4 = wb.create_sheet("Ringkasan")

summary = [
    ["LAPORAN ANALISIS DATA SANTRI - ANDI PRESENSI", ""],
    ["", ""],
    ["Tanggal Analisis:", "21 Agustus 2026"],
    ["", ""],
    ["STATISTIK TOTAL", ""],
    ["Total Santri:", len(all_students)],
    ["Total Keluarga Kakak Beradik Terverifikasi:", len(unique_families)],
    ["Total Siswa dalam Keluarga Sibling:", len(seen_nisns)],
    ["", ""],
    ["DATA DOBEL", ""],
    ["Jumlah NISN dobel (double entry):", sum(1 for s in nisn_groups.values() if len(s) > 1)],
    ["Jumlah record dobel:", sum(len(s) - 1 for s in nisn_groups.values() if len(s) > 1)],
    ["", ""],
    ["DATA PER KELAS", ""],
]

class_counts = defaultdict(int)
for s in all_students:
    class_counts[s['tingkat']] += 1

for kelas in ['Kelas 1', 'Kelas 2', 'Kelas 3', 'Kelas 4', 'Kelas 5', 'Kelas 6']:
    summary.append([kelas + ":", class_counts[kelas]])

summary.extend([
    ["", ""],
    ["VERIFIKASI STRATEGI", ""],
    ["Strategi 1 - Ayah + Ibu + Alamat:", f"{len(siblings1)} keluarga"],
    ["Strategi 2 - Ayah + Ibu + Telepon:", f"{len(siblings2)} keluarga"],
    ["Strategi 3 - Ayah + Ibu + Anak Ke:", f"{len(siblings3)} keluarga"],
    ["Strategi 4 - Alamat + Anak Ke:", f"{len(siblings4)} keluarga"],
])

for row_data in summary:
    ws4.append(row_data)

ws4.column_dimensions['A'].width = 40
ws4.column_dimensions['B'].width = 30

# Save
output_path = "DATA SANTRI/LAPORAN_KAKAK_BERADIK.xlsx"
wb.save(output_path)
print(f"Laporan disimpan: {output_path}")

# Print sample verified families
print("\n" + "="*80)
print("SAMPLE KELUARGA TERVERIFIKASI (10 TERBAIK)")
print("="*80)

for i, (key, students, strat) in enumerate(unique_families[:10]):
    print(f"\nKELUARGA {i+1} [{strat}]")
    print(f"  Ayah: {key[0] if isinstance(key, tuple) else 'N/A'}")
    print(f"  Ibu: {key[1] if isinstance(key, tuple) and len(key) > 1 else 'N/A'}")
    for s in students:
        tgl = str(s['tgl_lahir'])[:10] if pd.notna(s['tgl_lahir']) else '?'
        print(f"  - {s['nama']} [{s['tingkat']}] | TTL: {s['tempat']}, {tgl} | NISN: {s['nisn']}")
