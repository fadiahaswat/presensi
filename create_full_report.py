import pandas as pd
import numpy as np
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

files = [
    ('DATA SANTRI/25 06  2026_DATA SISWA KELAS 1 GABUNGAN 26-27ds.xlsx', 'Kelas 1'),
    ('DATA SANTRI/16072026_Terbaru DATA SISWA KELAS 2 GABUNGAN 26-27.xlsx', 'Kelas 2'),
    ('DATA SANTRI/18072026_Terbaru DATA SISWA KELAS 3 GABUNGAN 26-27ds.xlsx', 'Kelas 3'),
    ('DATA SANTRI/03 08 2026_DATA SISWA KELAS 4 GABUNGAN 26-27ds.xlsx', 'Kelas 4'),
    ('DATA SANTRI/20 07 2026_DATA SISWA KELAS 5 GABUNGAN 26-27ds.xlsx', 'Kelas 5'),
    ('DATA SANTRI/30 06 2026_DATA SISWA KELAS 6 GABUNGAN 26-27ds.xlsx', 'Kelas 6'),
]

all_students = []

for fpath, label in files:
    df = pd.read_excel(fpath)
    for i, row in df.iterrows():
        row_data = row.tolist()
        n_cols = len(row_data)

        if n_cols == 62:
            data = {
                'nama': row_data[4], 'nis': row_data[2], 'nisn': row_data[3],
                'jk': row_data[11], 'tempat': row_data[12], 'tgl_lahir': row_data[14],
                'ayah': row_data[27], 'ibu': row_data[34], 'alamat': row_data[18],
                'asal_sekolah': row_data[24], 'desa': row_data[20],
                'kecamatan': row_data[21], 'kota': row_data[22],
                'anak_ke': row_data[8], 'jml_saudara': row_data[9], 'telp_ayah': row_data[32],
                'tingkat': label, 'file': fpath,
            }
        else:
            data = {
                'nama': row_data[5], 'nis': row_data[3], 'nisn': row_data[4],
                'jk': row_data[12], 'tempat': row_data[13], 'tgl_lahir': row_data[15],
                'ayah': row_data[28], 'ibu': row_data[35], 'alamat': row_data[19],
                'asal_sekolah': row_data[25], 'desa': row_data[21],
                'kecamatan': row_data[22], 'kota': row_data[23],
                'anak_ke': row_data[9], 'jml_saudara': row_data[10], 'telp_ayah': row_data[33],
                'tingkat': label, 'file': fpath,
            }
        all_students.append(data)

def norm(s):
    if s is None or pd.isna(s): return ''
    return str(s).lower().strip()

# Styles
header_font = Font(bold=True, color="FFFFFF")
border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# Color fills
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
BLUE_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="F4B942", end_color="F4B942", fill_type="solid")
PURPLE_FILL = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")

wb = Workbook()

# =============================================
# SHEET 1: DATA DOBEL (NISN)
# =============================================
ws1 = wb.active
ws1.title = "1-Dobel-NISN"

ws1.cell(row=1, column=1, value="DATA DOBEL (NISN SAMA)").font = Font(bold=True, size=14)
ws1.cell(row=2, column=1, value="Data ini perlu dihapus salah satunya karena NISN-nya sama").font = Font(italic=True)

headers = ["NISN", "Nama", "Kelas", "JK", "Tempat Lahir", "Tanggal Lahir", "Ayah", "Ibu", "Aksi"]
for col, h in enumerate(headers, 1):
    cell = ws1.cell(row=4, column=col, value=h)
    cell.font = header_font
    cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    cell.border = border

nisn_groups = defaultdict(list)
for s in all_students:
    nisn = str(s['nisn']) if s['nisn'] and not pd.isna(s['nisn']) else ''
    if nisn:
        nisn_groups[nisn].append(s)

row_num = 5
for nisn, students in nisn_groups.items():
    if len(students) > 1:
        for s in students:
            tgl = str(s['tgl_lahir'])[:10] if pd.notna(s['tgl_lahir']) else ''
            data = [nisn, s['nama'], s['tingkat'], s['jk'], s['tempat'], tgl, s['ayah'], s['ibu'], 'HAPUS']
            for col, val in enumerate(data, 1):
                cell = ws1.cell(row=row_num, column=col, value=val)
                cell.border = border
                cell.fill = RED_FILL
            row_num += 1

for i, w in enumerate([15, 35, 10, 10, 20, 15, 30, 30, 15], 1):
    ws1.column_dimensions[chr(64+i)].width = w

# =============================================
# SHEET 2: ALAMAT SAMA ORANG TUA BERBEDA
# =============================================
ws2 = wb.create_sheet("2-Konfirmasi-Alamat")

ws2.cell(row=1, column=1, value="ALAMAT SAMA TAPI ORANG TUA BERBEDA").font = Font(bold=True, size=14)
ws2.cell(row=2, column=1, value="Kemungkinan: Salah data, Ayah berbeda (ayah+tiri?), atau Typo nama").font = Font(italic=True)

headers2 = ["Alamat", "Nama Santri", "NISN", "Kelas", "Nama Ayah", "Status", "Keterangan"]
for col, h in enumerate(headers2, 1):
    cell = ws2.cell(row=4, column=col, value=h)
    cell.font = header_font
    cell.fill = BLUE_FILL
    cell.border = border

alamat_groups = defaultdict(list)
for s in all_students:
    alamat = norm(s['alamat'])[:50] if s['alamat'] else ''
    if alamat and alamat != 'nan':
        alamat_groups[alamat].append(s)

row_num = 5
for alamat, students in alamat_groups.items():
    if len(students) >= 2:
        ayahs = set(norm(s['ayah']) for s in students if s['ayah'])
        if len(ayahs) > 1:
            # Check if ayah names are similar (typo)
            ayahs_list = list(ayahs)
            similar_ayahs = False
            for i, a1 in enumerate(ayahs_list):
                for a2 in ayahs_list[i+1:]:
                    if a1[:10] == a2[:10] or a1[-10:] == a2[-10:]:
                        similar_ayahs = True
                        break

            for s in students:
                if similar_ayahs:
                    status = "Kemungkinan TYP0"
                    fill = YELLOW_FILL
                else:
                    status = "Ayah Tiri?"
                    fill = ORANGE_FILL

                data = [alamat[:60], s['nama'], s['nisn'], s['tingkat'], s['ayah'], status, '']
                for col, val in enumerate(data, 1):
                    cell = ws2.cell(row=row_num, column=col, value=val)
                    cell.border = border
                    cell.fill = fill
                row_num += 1

for i, w in enumerate([60, 35, 15, 10, 30, 20, 30], 1):
    ws2.column_dimensions[chr(64+i)].width = w

# =============================================
# SHEET 3: TYP0 NAMA AYAH
# =============================================
ws3 = wb.create_sheet("3-Typo-Ayah")

ws3.cell(row=1, column=1, value="POTENSI TYPO NAMA AYAH").font = Font(bold=True, size=14)
ws3.cell(row=2, column=1, value="Nama yang sangat mirip mungkin typo").font = Font(italic=True)

headers3 = ["Nama Ayah 1", "Nama Ayah 2", "Kemiripan", "Aksi"]
for col, h in enumerate(headers3, 1):
    cell = ws3.cell(row=4, column=col, value=h)
    cell.font = header_font
    cell.fill = PURPLE_FILL
    cell.border = border

ayah_groups = defaultdict(list)
for s in all_students:
    ayah = norm(s['ayah'])
    if ayah and ayah != 'nan':
        ayah_groups[ayah].append(s)

# Find similar ayah names
similar_pairs = set()
all_ayahs = list(ayah_groups.keys())
for i, a1 in enumerate(all_ayahs):
    for a2 in all_ayahs[i+1:]:
        if len(a1) >= 5 and len(a2) >= 5:
            # Check first 10 chars
            if a1[:10] == a2[:10]:
                similar_pairs.add((a1, a2))
            # Check if only 1-2 chars different
            diff_count = sum(1 for c1, c2 in zip(a1, a2) if c1 != c2)
            if diff_count <= 3 and abs(len(a1) - len(a2)) <= 2:
                similar_pairs.add((a1, a2))

row_num = 5
for a1, a2 in sorted(similar_pairs):
    data = [a1.title(), a2.title(), "Kemungkinan Typo", "Cek Manual"]
    for col, val in enumerate(data, 1):
        cell = ws3.cell(row=row_num, column=col, value=val)
        cell.border = border
        cell.fill = YELLOW_FILL
    row_num += 1

for i, w in enumerate([35, 35, 20, 15], 1):
    ws3.column_dimensions[chr(64+i)].width = w

# =============================================
# SHEET 4: DATA TIDAK LENGKAP
# =============================================
ws4 = wb.create_sheet("4-Tidak-Lengkap")

ws4.cell(row=1, column=1, value="DATA YANG TIDAK LENGKAP").font = Font(bold=True, size=14)

headers4 = ["Nama", "NISN", "Kelas", "Field Kosong", "Nilai"]
for col, h in enumerate(headers4, 1):
    cell = ws4.cell(row=3, column=col, value=h)
    cell.font = header_font
    cell.fill = ORANGE_FILL
    cell.border = border

row_num = 4
for s in all_students:
    issues = []
    if not s['kota'] or pd.isna(s['kota']) or str(s['kota']) == 'nan':
        issues.append(('Kota', s['kota']))
    if not s['kecamatan'] or pd.isna(s['kecamatan']) or str(s['kecamatan']) == 'nan':
        issues.append(('Kecamatan', s['kecamatan']))
    if not s['telp_ayah'] or pd.isna(s['telp_ayah']) or str(s['telp_ayah']) == 'nan':
        issues.append(('Telp Ayah', s['telp_ayah']))

    if issues:
        for field, val in issues:
            data = [s['nama'], s['nisn'], s['tingkat'], field, val]
            for col, v in enumerate(data, 1):
                cell = ws4.cell(row=row_num, column=col, value=v)
                cell.border = border
            row_num += 1

for i, w in enumerate([35, 15, 10, 15, 30], 1):
    ws4.column_dimensions[chr(64+i)].width = w

# =============================================
# SHEET 5: ANALISIS ANAK KE
# =============================================
ws5 = wb.create_sheet("5-Validasi-AnakKe")

ws5.cell(row=1, column=1, value="ANALISIS ANAK KE vs JUMLAH SAUDARA").font = Font(bold=True, size=14)
ws5.cell(row=2, column=1, value="Catatan: Jika 'Anak ke' > 'Jml Saudara', kemungkinan definisi berbeda").font = Font(italic=True, color="666666")

headers5 = ["Nama", "NISN", "Kelas", "Anak Ke", "Jml Saudara", "Status", "Kemungkinan"]
for col, h in enumerate(headers5, 1):
    cell = ws5.cell(row=4, column=col, value=h)
    cell.font = header_font
    cell.fill = PatternFill(start_color="538135", end_color="538135", fill_type="solid")
    cell.border = border

row_num = 5
invalid_count = 0
for s in all_students:
    if pd.notna(s['anak_ke']) and pd.notna(s['jml_saudara']):
        if s['anak_ke'] > s['jml_saudara']:
            # Maybe jml_saudara doesn't count themselves
            anak_ke_adjusted = s['anak_ke']
            jml_adjusted = s['jml_saudara'] + 1

            if anak_ke_adjusted == jml_adjusted:
                status = "OK (jml tdk diritung sndiri)"
                fill = GREEN_FILL
            else:
                status = "INVALID"
                fill = RED_FILL
                invalid_count += 1

            data = [s['nama'], s['nisn'], s['tingkat'], int(s['anak_ke']), int(s['jml_saudara']), status, '']
            for col, val in enumerate(data, 1):
                cell = ws5.cell(row=row_num, column=col, value=val)
                cell.border = border
                cell.fill = fill
            row_num += 1

for i, w in enumerate([35, 15, 10, 10, 12, 30, 30], 1):
    ws5.column_dimensions[chr(64+i)].width = w

# =============================================
# SHEET 6: KAKAK BERADIK TERVERIFIKASI
# =============================================
ws6 = wb.create_sheet("6-Siblings-Terverifikasi")

ws6.cell(row=1, column=1, value="KAKAK BERADIK TERVERIFIKASI (TRIANGULASI)").font = Font(bold=True, size=14)
ws6.cell(row=2, column=1, value="Diverifikasi berdasarkan minimal 2 kriteria: Ayah+Ibu+Alamat, Ayah+Ibu+Telp, atau Ayah+Ibu+AnakKe").font = Font(italic=True)

# Run triangulation again
groups1 = defaultdict(list)
for s in all_students:
    ayah = norm(s['ayah'])
    ibu = norm(s['ibu'])
    alamat = norm(s['alamat'])[:60] if s['alamat'] else ''
    if ayah and ibu and alamat and ayah != 'nan' and ibu != 'nan' and alamat != 'nan':
        groups1[(ayah, ibu, alamat)].append(s)
siblings1 = {k: v for k, v in groups1.items() if len(v) > 1}

groups2 = defaultdict(list)
for s in all_students:
    ayah = norm(s['ayah'])
    ibu = norm(s['ibu'])
    telp = str(s['telp_ayah'])[:12] if s['telp_ayah'] else ''
    if ayah and ibu and telp and ayah != 'nan' and ibu != 'nan' and len(telp) >= 10:
        groups2[(ayah, ibu, telp)].append(s)
siblings2 = {k: v for k, v in groups2.items() if len(v) > 1}

all_strategies = [(siblings1, 'Ayah+Ibu+Alamat'), (siblings2, 'Ayah+Ibu+Telp')]
student_matches = defaultdict(set)
for groups, strat_name in all_strategies:
    for key, students in groups.items():
        for s in students:
            nisn = str(s['nisn']) if pd.notna(s['nisn']) else str(id(s))
            student_matches[nisn].add(strat_name)

verified_nisns = {n for n, strats in student_matches.items() if len(strats) >= 2}

headers6 = ["No", "Nama Santri", "NISN", "Kelas", "JK", "Anak ke", "TTL", "Nama Ayah", "Nama Ibu", "Alamat", "Kota", "Telp", "Verifikasi", "Confidence"]
for col, h in enumerate(headers6, 1):
    cell = ws6.cell(row=4, column=col, value=h)
    cell.font = header_font
    cell.fill = GREEN_FILL
    cell.border = border

row_num = 5
fam_num = 0
seen_nisns = set()

for groups, strat_name in all_strategies:
    for key, students in groups.items():
        verified = [s for s in students if str(s['nisn']) in verified_nisns]
        if len(verified) >= 2:
            fam_num += 1
            for s in students:
                if str(s['nisn']) not in seen_nisns:
                    seen_nisns.add(str(s['nisn']))
                    tgl = str(s['tgl_lahir'])[:10] if pd.notna(s['tgl_lahir']) else ''
                    anak = int(s['anak_ke']) if pd.notna(s['anak_ke']) else ''

                    nisn = str(s['nisn']) if pd.notna(s['nisn']) else ''
                    conf = "HIGH" if len(student_matches.get(nisn, set())) >= 2 else "MEDIUM"

                    data = [fam_num, s['nama'], s['nisn'], s['tingkat'], s['jk'], anak,
                           f"{s['tempat']}, {tgl}" if s['tempat'] else tgl,
                           s['ayah'], s['ibu'], str(s['alamat'])[:50] if s['alamat'] else '',
                           s['kota'], str(s['telp_ayah'])[:15] if s['telp_ayah'] else '',
                           strat_name, conf]

                    for col, val in enumerate(data, 1):
                        cell = ws6.cell(row=row_num, column=col, value=val)
                        cell.border = border
                        cell.fill = GREEN_FILL if conf == "HIGH" else PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
                    row_num += 1

for i, w in enumerate([6, 30, 12, 8, 8, 8, 25, 25, 25, 50, 15, 15, 20, 10], 1):
    ws6.column_dimensions[chr(64+i)].width = w

# =============================================
# SHEET 7: RINGKASAN
# =============================================
ws7 = wb.create_sheet("7-Ringkasan")

summary = [
    ["LAPORAN ANALISIS DATA SANTRI", "", ""],
    ["Generated: 21 Agustus 2026", "", ""],
    ["", "", ""],
    ["STATISTIK", "", ""],
    ["Total Santri", len(all_students), ""],
    ["", "", ""],
    ["MASALAH DITEMUKAN", "JUMLAH", "SEVERITY"],
    ["NISN Dobel (double entry)", sum(1 for s in nisn_groups.values() if len(s) > 1), "CRITICAL"],
    ["Alamat sama, ortu berbeda", len([a for a, s in alamat_groups.items() if len(s) >= 2 and len(set(norm(x['ayah']) for x in s)) > 1]), "HIGH"],
    ["Potensi Typo Nama Ayah", len(similar_pairs), "MEDIUM"],
    ["Data tidak lengkap", len([s for s in all_students if not s['kota'] or pd.isna(s['kota'])]), "LOW"],
    ["Invalid: Anak ke > Jml Saudara", invalid_count, "INFO"],
    ["", "", ""],
    ["KAKAK BERADIK", len(verified_nisns), "siswa"],
    ["Total Keluarga Terverifikasi", fam_num, "keluarga"],
]

for row_data in summary:
    ws7.append(row_data)

ws7.column_dimensions['A'].width = 40
ws7.column_dimensions['B'].width = 15
ws7.column_dimensions['C'].width = 15

# Save
output_path = "DATA SANTRI/LAPORAN_ANALISIS-LENGKAP.xlsx"
wb.save(output_path)
print(f"Laporan lengkap disimpan: {output_path}")
print(f"\nSummary:")
print(f"  - NISN Dobel: {sum(1 for s in nisn_groups.values() if len(s) > 1)}")
print(f"  - Alamat konflik: {len([a for a, s in alamat_groups.items() if len(s) >= 2 and len(set(norm(x['ayah']) for x in s)) > 1])}")
print(f"  - Typo Ayah: {len(similar_pairs)}")
print(f"  - Keluarga Terverifikasi: {fam_num}")
